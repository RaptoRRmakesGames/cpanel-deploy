from object_management import app, auth, admin, get_next_seven_days

from flask import (
    Flask,
    render_template,
    get_flashed_messages,
    flash,
    session,
    request,
    redirect,
    url_for,
    jsonify,
    make_response,
    send_file,
)
from datetime import timedelta, datetime

import db

import json

import time

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill

from io import BytesIO

import pandas as pd


@app.route("/table")
@app.route("/table/<week>")
def table(week=None):
    if not auth():
        return redirect(url_for("login_page"))

    match request.method:

        case "GET":

            if week == None:

                group = db.KitchenGroup(db.get_current_week())

                day_start = group.week.split("-")[2].strip()
                month_start = group.week.split("-")[1].strip()
                year_start = group.week.split("-")[0].split("(")[1].strip()

            else:

                year_start = week.split("_")[2]
                month_start = week.split("_")[1]
                day_start = week.split("_")[0]

                year_end = week.split("_")[5]
                month_end = week.split("_")[4]
                day_end = week.split("_")[3]

                formated_week = f"({year_start}-{month_start}-{day_start} - {year_end}-{month_end}-{day_end})"

                group = db.KitchenGroup(formated_week)

            dates = get_next_seven_days(f"{day_start}/{month_start}/{year_start}")

            all_employees = group.get_unplaced_employees()

            all_programs = db.get_all_programs()

            all_departments = db.Department.get_all_departments()

            weeks = db.get_next_weeks(4)
            selected_week = db.get_current_week() if week == None else formated_week

            all_weeks_saved = db.KitchenGroup.get_saved_weeks()

            todays_week = db.get_current_week()

            split_days = group.get_split_days()

            print(split_days)

            new_week_message = (
                ""
                if week == None
                else (
                    "Week Successfully Created. Make Sure to Save!"
                    if not group.saved
                    else ""
                )
            )

            return render_template(
                "table.html",
                session=session,
                group=group,
                all_employees=all_employees,
                all_programs=all_programs,
                all_departments=all_departments,
                weeks=weeks,
                selected_week=selected_week,
                all_weeks_saved=all_weeks_saved,
                new_week_message=new_week_message,
                todays_week=todays_week,
                dates=dates,
                split_days=split_days,
            )


@app.route("/see_week/<d_start>_<m_start>_<y_start>_<d_end>_<m_end>_<y_end>")
def see_week(d_start, m_start, y_start, d_end, m_end, y_end):
    if not auth():
        return redirect(url_for("login_page"))

    return f"{d_start}/{m_start}/{y_start} - {d_end}/{m_end}/{y_end}"


@app.route("/save_table_excel", methods=["GET", "POST"])
def save_table_excel():
    if not auth():
        return redirect(url_for("login_page"))

    data = dict(request.json)
    rows = []
    week_name = ""
    for location, events in data.items():
        if location == "week":
            week_name = events.replace("(", "").replace(")", "")
            continue
        rows.append({"Name": location})
        for event, details in events.items():
            rows.append({"Name": event})
            if details:  # Check if the event has details
                for detail in details:
                    employee_name = detail[0]
                    for i, days_info in enumerate(detail[1]):
                        row = {
                            "Name": f"{employee_name if i == 0 else f'Split {i}'}",
                            "Monday": days_info.get("monday", ["", ""])[0],
                            "Change M": days_info.get("monday", ["", ""])[1],
                            "Tuesday": days_info.get("tuesday", ["", ""])[0],
                            "Change T": days_info.get("tuesday", ["", ""])[1],
                            "Wednesday": days_info.get("wednesday", ["", ""])[0],
                            "Change W": days_info.get("wednesday", ["", ""])[1],
                            "Thursday": days_info.get("thursday", ["", ""])[0],
                            "Change Tu": days_info.get("thursday", ["", ""])[1],
                            "Friday": days_info.get("friday", ["", ""])[0],
                            "Change F": days_info.get("friday", ["", ""])[1],
                            "Saturday": days_info.get("saturday", ["", ""])[0],
                            "Change Sa": days_info.get("saturday", ["", ""])[1],
                            "Sunday": days_info.get("sunday", ["", ""])[0],
                            "Change Su": days_info.get("sunday", ["", ""])[1],
                        }
                        rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Save to an Excel buffer with openpyxl engine for styling
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    # Load the workbook and the first sheet for styling
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb.active

    # Query department and kitchen names from the database
    department_names = db.execute(
        "SELECT name FROM sub_department WHERE user_id=%s", [session["user_id"]]
    )
    kitchens_names = db.execute(
        "SELECT name FROM big_kitchens WHERE user_id=%s", [session["user_id"]]
    )
    day_names = [
        "Name",
        "Monday", "Change M",
        "Tuesday", "Change T",
        "Wednesday", "Change W",
        "Thursday", "Change Tu",
        "Friday", "Change F",
        "Saturday", "Change Sa",
        "Sunday", "Change Su"
    ]

    department_names = [f[0] for f in department_names]
    kitchens_names = [f[0] for f in kitchens_names]

    print(department_names, kitchens_names)

    # Define thicker borders
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    # Define fill colors for departments, kitchens, and days
    department_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    kitchen_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")    # Yellow
    day_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")        # Red

    # Apply thicker borders, center alignment, and conditional formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thick_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply color fill based on cell value
            if cell.value in department_names:
                cell.fill = department_fill
            elif cell.value in kitchens_names:
                cell.fill = kitchen_fill
            elif cell.value in day_names:
                cell.fill = day_fill

    # Adjust column width for better visibility
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the letter of the column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding

    # Save the styled Excel to a new buffer
    styled_excel_buffer = BytesIO()
    wb.save(styled_excel_buffer)
    styled_excel_buffer.seek(0)

    # Send the response with styled Excel
    response = make_response(styled_excel_buffer.read())
    response.headers["Content-Disposition"] = (
        f"attachment; filename=report_{week_name}.xlsx"
    )
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return response