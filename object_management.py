from create_app import app, auth, admin, get_next_seven_days

from flask import (
    Flask,
    render_template,
    get_flashed_messages,
    flash,
    session,
    request,
    redirect,
    url_for,
    jsonify,
    make_response,
    send_file,
    Response
)
from datetime import timedelta, datetime

import db

import json

import time

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from io import BytesIO

import pandas as pd 

import calendar

from collections import defaultdict

@app.route("/")
def index():
    if not auth():
        return redirect(url_for("login_page"))

    return render_template(
        "index.html",
        session=session,
    )



@app.route("/add_kitchen", methods=["GET", "POST"])
def add_kitchen():

    if not auth():
        return redirect(url_for("login_page"))

    match request.method:
        case "GET":
            return render_template(
                "add_kitchen.html",
                session=session,
                departments=db.get_all_departments(),
            )

        case "POST":

            name, deps = request.form.get("name"), []
            for key in request.form:
                if key == "name" or key == "submit":
                    continue
                deps.append(request.form.get(key))

            db.Kitchen.create_kitchen(name, deps)

            flash("Created Kitchen Successfully!")

            return redirect(url_for("edit_objects"))


@app.route("/add_department", methods=["GET", "POST"])
def add_dep():

    if not auth():
        return redirect(url_for("login_page"))
    match request.method:
        case "GET":
            return render_template(
                "add_deps.html",
                session=session,
            )

        case "POST":

            db.create_dept(request.form.get("name"))

            flash("Created Department Successfully!")

            return redirect(url_for("edit_objects"))


@app.route("/add_title", methods=["GET", "POST"])
def add_title():
    if not auth():
        return redirect(url_for("login_page"))
    match request.method:
        case "GET":
            return render_template(
                "add_title.html",
                session=session,
            )

        case "POST":

            db.create_title(request.form.get("name"))

            flash("Created Title Successfully!")

            return redirect(url_for("edit_objects"))


@app.route("/save_schedule", methods=["POST"])
def save_schedule():
    if not auth():
        return redirect(url_for("login_page"))
    data = request.json

    schedule = db.KitchenGroup(data["week"])

    schedule.load_schedule(data)

    schedule.save_schedule()

    return jsonify(
        {
            "status": "success",
        }
    )


@app.route("/add_employee", methods=["GET", "POST"])
def add_employee():
    if not auth():
        return redirect(url_for("login_page"))
    match request.method:

        case "GET":

            all_departments = db.Department.get_all_departments()
            all_programs = db.get_all_programs()
            all_titles = db.get_all_titles()

            return render_template(
                "add_employee.html",
                session=session,
                departments=all_departments,
                programs=all_programs,
                titles=all_titles,
            )

        case "POST":

            name, title, def_dep, days_per_week,salary, salary13, salary14, leave, gesy, pro_fund, guild, time,code= (
                request.form.get("name").lower().capitalize(),
                request.form.get("title"),
                request.form.get("def_dep"),
                0 if request.form.get("days_per_week") is None else request.form.get("days_per_week"),
                0.0 if request.form.get("salary") is None or request.form.get("salary") ==''  else request.form.get("salary"),
                0 if request.form.get("13salary") is None else 1 ,
                0 if request.form.get("14salary")is None else 1 ,
                0 if request.form.get("leave")is None else 1 ,
                0 if request.form.get("gesy") is None or request.form.get("gesy") =='' else request.form.get("gesy"),
                0 if request.form.get("provident_fund")is None or request.form.get("provident_fund") =='' else request.form.get("provident_fund"),
                0 if request.form.get("guild") is None or request.form.get("guild") =='' else request.form.get("guild"),
                request.form.get('time'),
                request.form.get('code'),
                
                
            )
            print('code in add: ' , code)
            
            name_parts = name.split(" ")
            name=''
            for part in name_parts:
                name += part.capitalize() + ' '
            print(name, 'name')
            employee = db.Employee.create_employee(
                name, title, def_dep, salary, days_per_week, salary13, salary14, gesy, pro_fund, guild, leave, time,code )

            dbe, c = db.connect()

            c.execute(
                "SELECT * FROM schedules WHERE user_id=%s", [session["user_id"]]
            )
            kitch, dep = employee.prefered_dep
            for sched in c.fetchall():

                ide = sched[0]
                week = sched[1]
                js = json.loads(sched[2].replace("'", '"'))

                program = db.get_random_program2()

                for kitchen in js:

                    if kitchen == kitch:

                        for department in js[kitchen]:

                            if department == dep:

                                js[kitchen][department].append(
                                    [
                                        employee.id,
                                        [
                                            {
                                                "monday": [program, ""],
                                                "tuesday": [program, ""],
                                                "wednesday": [program, ""],
                                                "thursday": [program, ""],
                                                "friday": [program, ""],
                                                "saturday": [program, ""],
                                                "sunday": [program, ""],
                                            }
                                        ],
                                    ]
                                )

                flash(f"Added into `{week}` Schedule")
                c.execute(
                    "UPDATE schedules SET schedule_json =%s WHERE id=%s", [str(js), ide]
                )

                dbe.commit()

            flash("Added Employee Successfully!")
            return redirect(url_for("edit_objects"))
        

    


@app.route("/add_program", methods=["GET", "POST"])
def create_program():
    if not auth():
        return redirect(url_for("login_page"))
    match request.method:
        case "GET":

            return render_template(
                "add_program.html",
                session=session,
            )

        case "POST":

            start = str(request.form.get("start"))
            end = str(request.form.get("end"))
            name = f'{start}-{end}'
            print(start, end)

            db.add_program(name)

            flash("Program Created")

            return redirect(url_for("edit_objects"))


@app.route("/edit_objects")
def edit_objects():
    if not auth():
        return redirect(url_for("login_page"))
    if session['user_hotel_owner']:
        return redirect(url_for('index'))
    all_kitchens = db.Kitchen.get_all_kitchens(True)
    all_departments = db.get_all_departments(True)

    employee_departments = db.Department.get_all_departments()
    all_employees = db.Employee.get_all_employees()
    all_programs = db.get_all_programs(True)
    all_titles = db.get_all_titles(True)
    
    print(len(all_employees), 'all_employees')

    dbe, c = db.connect()

    c.execute("SELECT * FROM schedules WHERE user_id=%s", [session["user_id"]])
    all_schedules = []

    for thing in c.fetchall():

        week_name = thing[1]

        day_start = week_name.split("-")[2].strip()
        day_end = week_name.split("-")[5].split(")")[0]

        month_start = week_name.split("-")[1]
        month_end = week_name.split("-")[4]

        year_start = week_name.split("-")[0].split("(")[1]
        year_end = week_name.split("-")[3].strip()

        url_time = (
            f"{day_start}_{month_start}_{year_start}_{day_end}_{month_end}_{year_end}"
        )

        all_schedules.append(
            [
                thing[0],
                thing[1],
                thing[2],
                url_time,
            ]
        )

    return render_template(
        "edit_objects.html",
        session=session,
        kitchens=all_kitchens,
        departments=all_departments,
        employees=all_employees,
        programs=all_programs,
        titles=all_titles,
        employee_departments=employee_departments,
        schedules=all_schedules,
    )


@app.route("/delete/table/<ide>")
def delete_table(ide):

    dbe, c = db.connect()

    c.execute("DELETE FROM schedules WHERE id=%s", [ide])

    dbe.commit()

    flash(f"Deleted Schedule ID-`{ide}`!")

    return redirect(url_for("edit_objects"))


@app.route("/edit/kitchen/<kitchen>", methods=["GET", "POST"])
def edit_kitchen(kitchen):
    if not auth():
        return redirect(url_for("login_page"))
    match request.method:
        case "GET":
            return render_template(
                "add_kitchen.html",
                session=session,
                departments=db.get_all_departments(),
                edit=True,
                kitchen=db.Kitchen(db.Kitchen.get_id_from_name(kitchen)),
            )

        case "POST":

            name, deps = request.form.get("name"), []
            for key in request.form:
                if key == "name" or key == "submit" or key == "id":
                    continue
                deps.append(request.form.get(key))

            k = db.Kitchen(db.Kitchen.get_id_from_name(kitchen))

            db_obj, c = db.connect()
            if k.name != name:

                c.execute("SELECT id, default_dep, name FROM employees")

                emps = c.fetchall()

                for emp in emps:

                    empid = emp[0]
                    emp_def_kitch, emp_def_dep = emp[1].split(" - ")
                    empname = emp[2]

                    if emp_def_kitch == k.name:
                        c.execute(
                            "UPDATE employees SET default_dep=%s WHERE id=%s",
                            [f"{name} - {emp_def_dep}", empid],
                        )
                        db_obj.commit()

                        flash(f"Updated Employee `{empname}`")

            c.execute(
                "SELECT id, week, schedule_json FROM schedules WHERE user_id=%s ",
                [session["user_id"]],
            )

            f = c.fetchall()

            for sched in f:

                week_id = sched[0]
                week_name = sched[1]
                sched_json = json.loads(sched[2].replace("'", '"'))
                new_js = {}
                for kitchen_key in sched_json:

                    if kitchen_key == k.name:
                        new_js[name] = sched_json[kitchen_key]
                    else:

                        new_js[kitchen_key] = sched_json[kitchen_key]

                c.execute(
                    "UPDATE schedules SET schedule_json=%s WHERE id=%s",
                    [str(new_js), week_id],
                )
                db_obj.commit()

                flash(f"Week `{week_name}` updated for Kitchen name change")

            k.update(name, deps)

            flash("Updated Kitchen Successfully!")

            return redirect(url_for("edit_objects"))


@app.route("/edit/department/<dep>", methods=["GET", "POST"])
def edit_department(dep):
    if not auth():
        return redirect(url_for("login_page"))
    match request.method:
        case "GET":
            return render_template(
                "add_deps.html",
                session=session,
                departments=db.get_all_departments(),
                edit=True,
                department=db.Department(db.Department.get_id_from_name(dep)),
            )

        case "POST":

            name = request.form.get("name")

            old_dep = db.Department(db.Department.get_id_from_name(dep))

            db_obj, c = db.connect()

            c.execute("SELECT id, default_dep,name FROM employees")

            emps = c.fetchall()

            for emp in emps:

                empid = emp[0]

                empname = emp[2]

                def_dep_kitch, def_dep = emp[1].split(" - ")

                if old_dep.name == def_dep:
                    c.execute(
                        "UPDATE employees SET default_dep=%s WHERE id=%s",
                        [f"{def_dep_kitch} - {name}", empid],
                    )

                    db_obj.commit()
                    # flash(f"Updated Department in Employee `{empname}`")

            c.execute(
                "SELECT id, week, schedule_json FROM schedules WHERE user_id=%s ",
                [session["user_id"]],
            )

            f = c.fetchall()

            for sched in f:

                week_id = sched[0]
                week_name = sched[1]
                sched_json = json.loads(sched[2].replace("'", '"'))
                new_js = {}
                for kitchen_key in sched_json:
                    new_js[kitchen_key] = {}
                    for dep_key in sched_json[kitchen_key]:
                        if dep_key == old_dep.name:

                            new_js[kitchen_key][name] = []

                            for ls in sched_json[kitchen_key][dep_key]:
                                new_js[kitchen_key][name].append(ls)

                        else:

                            new_js[kitchen_key][dep_key] = sched_json[kitchen_key][
                                dep_key
                            ]

                c.execute(
                    "UPDATE schedules SET schedule_json=%s WHERE id=%s",
                    [str(new_js), week_id],
                )
                db_obj.commit()

                flash(f"Week `{week_name}` updated for department name change")

            old_dep.update(name)

            flash("Updated Department Successfully!")

            return redirect(url_for("edit_objects"))


@app.route("/edit/employee/<emp>", methods=["GET", "POST"])
def edit_demployee(emp):
    if not auth():
        return redirect(url_for("login_page"))
    match request.method:
        case "GET":

            return render_template(
                "add_employee.html",
                session=session,
                departments=db.Department.get_all_departments(),
                edit=True,
                employee=db.Employee(db.Employee.get_id_by_name(emp)),
                titles=db.get_all_titles(),
                programs=db.get_all_programs(),
            )

        case "POST":

            name = request.form.get("name").lower().capitalize()
            title = request.form.get("title")
            pref_dep = request.form.get("def_dep")
            name, title, def_dep, days_per_week,salary, salary13, salary14, leave, gesy, pro_fund, guild, time,code= (
                request.form.get("name"),
                request.form.get("title"),
                request.form.get("def_dep"),
                0 if request.form.get("days_per_week") is None else request.form.get("days_per_week"),
                0.0 if request.form.get("salary") is None or request.form.get("salary") ==''  else request.form.get("salary"),
                0 if request.form.get("13salary") is None else 1 ,
                0 if request.form.get("14salary")is None else 1 ,
                0 if request.form.get("leave")is None else 1 ,
                0 if request.form.get("gesy") is None or request.form.get("gesy") =='' else request.form.get("gesy"),
                0 if request.form.get("provident_fund")is None or request.form.get("provident_fund") =='' else request.form.get("provident_fund"),
                0 if request.form.get("guild") is None or request.form.get("guild") =='' else request.form.get("guild"),
                request.form.get('time'), request.form.get('code')
                
            )
            
            print('fwaeh', days_per_week)

            db.Employee(db.Employee.get_id_by_name(emp)).update(name, title, pref_dep, salary, days_per_week, salary13, salary14, leave, gesy, pro_fund, guild, time, code)

            flash("Updated Department Successfully!")

            return redirect(url_for("edit_objects"))


@app.route("/delete/program/<program>")
def delete_program(program):
    if not auth():
        return redirect(url_for("login_page"))

    db_obj, c = db.connect()

    program = db.remove_from_string(program)

    c.execute(
        "SELECT schedule_json FROM schedules WHERE user_id=%s", [session["user_id"]]
    )
    remove = True
    for schedule in c.fetchall():

        if program.strip() in schedule[0]:
            remove = False
            break

    if remove:
        flash(f"`{program}` Deleted Successfully")

        c.execute("DELETE FROM programs WHERE name=%s AND user_id != -1", [program])
        db_obj.commit()
    else:
        flash(f"`{program}` Could not be deleted because it is used in a Schedule!.")
    return redirect(url_for("edit_objects"))


@app.route("/delete/title/<title>")
def delete_title(title):
    if not auth():
        return redirect(url_for("login_page"))

    db_obj, c = db.connect()

    title = db.remove_from_string(title).strip()

    c.execute(
        "SELECT id, name, title FROM employees WHERE user_id=%s", [session["user_id"]]
    )

    remove = True
    for emp in c.fetchall():

        if remove == False:
            break

        remove = not emp[2].strip() == db.remove_from_string(title).strip()

    if remove:
        title = db.remove_from_string(title)
        flash(f"`{title}` Deleted Successfully")
        c.execute("DELETE FROM titles WHERE name=%s ", [title])
        db_obj.commit()
    else:
        flash(f"`{title}` couldnt be deleted because its used in an Employee!")

    return redirect(url_for("edit_objects"))


@app.route("/delete/kitchen/<kitchen>")
def delete_kitchen(kitchen):
    if not auth():
        return redirect(url_for("login_page"))

    kitchen = db.remove_from_string(kitchen).strip()

    dbe, c = db.connect()

    c.execute("SELECT id FROM employees")
    remove = True
    for emp in c.fetchall():

        if not remove:
            break

        empl = db.Employee(emp[0])

        try:
            remove = not db.remove_from_string(kitchen).strip() in empl.prefered_dep_str
        except AttributeError:
            remove = True

    if remove:
        db_obj, c = db.connect()
        kitchen = db.remove_from_string(kitchen)
        flash(f"`{kitchen}` Deleted Successfully")
        c.execute("DELETE FROM big_kitchens WHERE name=%s ", [kitchen])
        db_obj.commit()

    else:
        flash(f"`{kitchen}` could not be deleted because its used in an Employee!")

    return redirect(url_for("edit_objects"))


@app.route("/delete/department/<department>")
def delete_department(department):
    if not auth():
        return redirect(url_for("login_page"))
    dbe, c = db.connect()

    department = db.remove_from_string(department).strip()

    c.execute("SELECT id FROM employees")
    remove = True
    for emp in c.fetchall():

        if not remove:
            break

        empl = db.Employee(emp[0])

        try:
            remove = not department.strip() in empl.prefered_dep_str
        except AttributeError:
            remove = True

    if remove:

        db_obj, c = db.connect()
        department = db.remove_from_string(department)
        flash(
            f"`{department}` Deleted Successfully. Make sure to edit Employees that were assigned to that department!"
        )
        c.execute("DELETE FROM sub_department WHERE name=%s ", [department])
        db_obj.commit()

    else:
        flash(f"`{department}` could not be deleted because its used in an Employee!")

    return redirect(url_for("edit_objects"))


@app.route("/delete/employee/<employee>")
def delete_employee(employee):
    if not auth():
        return redirect(url_for("login_page"))

    emp_name = db.remove_from_string(employee).strip()

    db_obj, c = db.connect()

    empl = db.Employee(db.Employee.get_id_by_name(emp_name))

    if empl == None:
        flash("Error Deleting Employee")
        return redirect(url_for("edit_objects"))

    empl.copy_to_archive()
    c.execute("DELETE FROM employees WHERE id=%s ", [empl.id])
    flash(f"`{emp_name}` Deleted Successfully")
    db_obj.commit()

    return redirect(url_for("edit_objects"))

@app.route("/create_objects/<page>", methods=["GET", "POST"])
def object_startup(page):

    match request.method:

        case "GET":
            all_departments = db.get_all_departments()
            all_programs = db.get_all_programs()
            all_titles = db.get_all_titles()
            employee_departments = db.Department.get_all_departments()
            return render_template(
                "object_startup.html",
                session=session,
                title="Startup Objects",
                page=page,
                object_name={
                    "1": "department",
                    "2": "kitchen",
                    "3": "program",
                    "4": "title",
                    "5": "employee",
                },
                departments=all_departments,
                programs=all_programs,
                titles=all_titles,
                employee_departments=employee_departments,
                edit=False,
            )

        case "POST":

            match request.form.get("add_type"):

                case "department":

                    db.create_dept(request.form.get("name"))

                    flash("Department Created Successfully!")

                    return redirect(url_for("object_startup", page=page))

                case "kitchen":

                    name, deps = request.form.get("name"), []
                    for key in request.form:
                        if key == "name" or key == "submit" or key == "add_type":
                            continue
                        deps.append(request.form.get(key))

                    db.Kitchen.create_kitchen(name, deps)

                    flash("Created Kitchen Successfully!")
                    return redirect(url_for("object_startup", page=page))

                case "program":
                    start = str(request.form.get("start"))
                    end = str(request.form.get("end"))
                    name = f'{start} - {end}'
                    print(start, end)

                    db.add_program(name)

                    flash("Program Created Successfully!")
                    return redirect(url_for("object_startup", page=page))

                case "title":

                    db.create_title(request.form.get("name"))

                    flash("Title Created Successfully!")
                    return redirect(url_for("object_startup", page=page))

                case "employee":

                    name, title, def_dep = (
                        request.form.get("name"),
                        request.form.get("title"),
                        request.form.get("def_dep"),
                    )

                    db.Employee.create_employee(name, title, def_dep)

                    flash("Employee Created Successfully!")
                    return redirect(url_for("object_startup", page=page))

                case _:

                    flash(f"Page #{page} is not valid!")

                    return redirect(url_for("object_startup", page=page))


@app.route("/save_department", methods=["POST"])
def save_department():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE sub_department SET name=%s WHERE id=%s",
            [data["new_name"], data["id"]],
        )

        old_dep = db.Department(data["id"])
        name = data["new_name"]

        c.execute("SELECT id, default_dep,name FROM employees")

        emps = c.fetchall()

        for emp in emps:

            empid = emp[0]

            empname = emp[2]

            def_dep_kitch, def_dep = emp[1].split(" - ")

            if old_dep.name == def_dep:
                c.execute(
                    "UPDATE employees SET default_dep=%s WHERE id=%s",
                    [f"{def_dep_kitch} - {name}", empid],
                )

                dbe.commit()
                # flash(f"Updated Department in Employee `{empname}`")

        c.execute(
            "SELECT id, week, schedule_json FROM schedules WHERE user_id=%s ",
            [session["user_id"]],
        )

        f = c.fetchall()

        for sched in f:

            week_id = sched[0]
            week_name = sched[1]
            sched_json = json.loads(sched[2].replace("'", '"'))
            new_js = {}
            for kitchen_key in sched_json:
                new_js[kitchen_key] = {}
                for dep_key in sched_json[kitchen_key]:
                    if dep_key == old_dep.name:

                        new_js[kitchen_key][name] = []

                        for ls in sched_json[kitchen_key][dep_key]:
                            new_js[kitchen_key][name].append(ls)

                    else:
                        new_js[kitchen_key][dep_key] = sched_json[kitchen_key][dep_key]

            c.execute(
                "UPDATE schedules SET schedule_json=%s WHERE id=%s",
                [str(new_js), week_id],
            )
            dbe.commit()

            # flash(f"Week `{week_name}` updated for department name change")

        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()


@app.route("/save_kitchen", methods=["POST"])
def save_kitchen():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE big_kitchens SET name=%s WHERE id=%s",
            [data["new_name"], data["id"]],
        )

        name = data["new_name"]
        k = db.Kitchen(data["id"])

        db_obj, c = db.connect()
        if k.name != name:

            c.execute(
                "SELECT id, default_dep, name FROM employees WHERE user_id=%s",
                [session["user_id"]],
            )

            emps = c.fetchall()

            for emp in emps:

                empid = emp[0]
                emp_def_kitch, emp_def_dep = emp[1].split(" - ")
                empname = emp[2]

                if emp_def_kitch == k.name:
                    c.execute(
                        "UPDATE employees SET default_dep=%s WHERE id=%s",
                        [f"{name} - {emp_def_dep}", empid],
                    )
                    db_obj.commit()

                    flash(f"Updated Employee `{empname}`")

        c.execute(
            "SELECT id, week, schedule_json FROM schedules WHERE user_id=%s ",
            [session["user_id"]],
        )

        f = c.fetchall()

        for sched in f:

            week_id = sched[0]
            week_name = sched[1]
            sched_json = json.loads(sched[2].replace("'", '"'))
            new_js = {}
            for kitchen_key in sched_json:

                if kitchen_key == k.name:
                    new_js[name] = sched_json[kitchen_key]
                else:

                    new_js[kitchen_key] = sched_json[kitchen_key]

            c.execute(
                "UPDATE schedules SET schedule_json=%s WHERE id=%s",
                [str(new_js), week_id],
            )
            db_obj.commit()

            # flash(f"Week `{week_name}` updated for Kitchen name change")

        dbe.commit()

        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()


@app.route("/save_program", methods=["POST"])
def save_program():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE programs SET name=%s WHERE id=%s AND user_id != %s", [data["new_name"], data["id"], -1]
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()
        
@app.route("/save_program_start", methods=["POST"])
def save_program_start():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute('SELECT * FROM programs WHERE id=%s', [data['id']])
        old_time = c.fetchall()[0][1]
        new_time = f"{data['new_name']}-{old_time.split('-')[1]}"
        c.execute('UPDATE programs SET name=%s WHERE id=%s ', [new_time, data['id']])
        print(new_time)
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()
@app.route("/save_program_end", methods=["POST"])
def save_program_end():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute('SELECT * FROM programs WHERE id=%s', [data['id']])
        old_time = c.fetchall()[0][1]
        new_time = f"{old_time.split('-')[0]}-{data['new_name']}"
        c.execute('UPDATE programs SET name=%s WHERE id=%s ', [new_time, data['id']])
        print(new_time)
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()



@app.route("/save_title", methods=["POST"])
def save_title():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE titles SET name=%s WHERE id=%s", [data["new_name"], data["id"]]
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()


@app.route("/save_emp_name", methods=["POST"])
def save_emp_name():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE employees SET name=%s WHERE id=%s", [data["new_name"], data["id"]]
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()
        
@app.route("/save_emp_code", methods=["POST"])
def save_emp_code():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE employees SET code=%s WHERE id=%s", [data["new_code"], data["id"]]
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()
        
@app.route("/save_emp_program", methods=["POST"])
def save_emp_time():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE employees SET pref_time=%s WHERE id=%s", [data["new_time"], data["id"]]
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()


@app.route("/save_emp_title", methods=["POST"])
def save_emp_title():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE employees SET title=%s WHERE id=%s", [data["new_name"], data["id"]]
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()


@app.route("/save_emp_pref_dep", methods=["POST"])
def save_emp_pref_dep():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.json
    print("Received data:", data)  # Debug statement to check received data

    try:
        dbe, c = db.connect()
        c.execute(
            "UPDATE employees SET default_dep=%s WHERE id=%s",
            [data["new_name"], data["id"]],
        )
        dbe.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        print("Error:", e)  # Debug statement to check the error
        return jsonify({"status": "error", "message": str(e)})
    finally:
        c.close()
        dbe.close()


def create_excel_employee():
    # set text
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    # set text
    ws['A1'] = 'Name'
    ws['B1'] = 'Title'
    ws['C1'] = 'Department'
    # set dropdown
    dropdown_options1 = db.Department.get_all_departments()
    dropdown_options2 = db.get_all_titles()
    for i in range(50):
        dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_options1)}"')
        ws.add_data_validation(dv)
        dv.add(ws[f"C{1+i}"])
        dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_options2)}"')
        ws.add_data_validation(dv)
        dv.add(ws[f"B{1+i}"])
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    
    excel_buffer.seek(0)
    
    excel_data = excel_buffer.read()
    
    response = make_response(excel_data)
    response.headers['Cache-Control'] = 'no-store'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Content-Disposition'] = 'attachment; filename=report.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@app.route('/get_excel_add_employee', methods=['POST'])
def get_excel_add_employee():
    try:
        return create_excel_employee()
    except Exception as e:
        return str(e)
    
@app.route('/change_kitchen_row', methods=['GET', 'POST'])
def save_kitchen_row():
    if not auth():
        return redirect(url_for("login_page"))
    
    match request.method:
        
        case 'GET':
            
            return render_template('kitchen_row.html', kitchens = db.Kitchen.get_all_kitchens(True))
        
        case 'POST':
            
            dbe,c = db.connect()
            
            for kitchen_id in request.form:
                
                row = request.form.get(kitchen_id)
                
                c.execute('UPDATE big_kitchens SET row=%s WHERE id=%s', [kitchen_id, row])
                
            dbe.commit()
            
            return redirect(url_for('save_kitchen_row'))

@app.route('/add_from_excel', methods=['GET', 'POST'])
def add_from_excel():
    
    match request.method:
        
        case 'GET':
            return render_template('add_from_excel.html')
        
        case 'POST':
            
            if 'file' not in request.files:
                flash('Please Choose a File')
                return redirect(url_for('add_from_excel'))
            
            fil = request.files['file']
            
            if fil.filename == '':
                flash('Please Choose a File')
                return redirect(url_for('add_from_excel'))
            try:
                df = pd.read_excel(fil)
                processed_data = []

                # Iterate over each row in the DataFrame
                for index, row in df.iterrows():
                    name = row['Name']
                    title = row['Title']
                    department = row['Department']
                    
                    # Check if the 'Name' field is not empty
                    if pd.notna(name) and name.strip() != "":
                        # Perform your logic here
                        # For demonstration, let's assume we are simply logging the row
                        # and creating a processed data structure
                        processed_row = {
                            "Name": name,
                            "Title": title,
                            "Department": department
                        }
                        processed_data.append(processed_row)
                        
                    for row in processed_data:
                        if db.Employee.create_employee(row['Name'], row['Title'], row['Department']) != False:  
                            flash('`' + row['Name'] +  '` Created!')
                        else:
                            flash(f'Name { row["Name"]} Already Exists')
                        
            except Exception as e:
                flash(f'Error: {e}')
                return redirect(url_for('add_from_excel'))
            
            flash('Successfully Created All Employees!')
            return redirect(url_for('add_from_excel'))
        
@app.route('/select_bednight_date', methods=['GET', 'POST'])
def select_bednight_date():
    if not auth():
        return redirect(url_for("login_page"))
    
    match request.method:
        
        case 'GET':
            return render_template('select_bednight_date.html', title='Select Bednight Date', owner = session['user_hotel_owner'])
        
        case 'POST':
            
            date_start = request.form.get('date_start')
            date_end = request.form.get('date_end')
            print(request.form.get('mode'))
            if request.form.get('mode') == 'add':
                return redirect(url_for('add_bednight',date_start = date_start, date_end = date_end))
            else:
                return redirect(url_for('manage_bednights',start_date = date_start, end_date = date_end))
        
@app.route('/add_bednight/<date_start>/<date_end>', methods=['GET',])
def add_bednight(date_start, date_end):
    if not auth():
        return redirect(url_for("login_page"))
    if not session['user_hotel_owner']:
        return redirect(url_for('index'))
    

    days = []

    start_date = datetime.strptime(date_start, "%Y-%m-%d")
    end_date = datetime.strptime(date_end, "%Y-%m-%d")

    current_date = start_date
    while current_date <= end_date:
        days.append(current_date.strftime("%Y-%m-%d"))
        current_date += timedelta(days=1)

    return render_template('add_bednight.html', title='Add Bednight',
                           date_start=date_start, date_end=date_end, days=days)
    
@app.route('/add_bednight', methods=['POST'])
def add_bednight_row():
    if not session['user_hotel_owner']:
        return redirect(url_for('index'))
    data = request.get_json()

    date = data.get('date')
    zero_six = data.get('zero_six')
    six_twelve = data.get('six_twelve')
    twelve_eighteen = data.get('twelve_eighteen')
    adults = data.get('adults')

    if None in (date, zero_six, six_twelve, twelve_eighteen, adults):
        return jsonify({"error": "Missing data"}), 400

    # Here, you would typically save this data to your database.
    # For demonstration, let's print the received data.
    print(f"Received data: {data}")
    
    
    
    dbe,c = db.connect()
    c.execute('SELECT id FROM bednights WHERE (user_id=%s OR user_id=%s)  AND date=%s', [session['user_id'],session['parent_id'], date])
    existing_bednight = c.fetchall()
    if len(existing_bednight) > 0:
        c.execute('UPDATE bednights SET zero_six=%s, six_twelve=%s, twelve_eighteen=%s, adults=%s WHERE id=%s',
                 [zero_six, six_twelve, twelve_eighteen, adults, existing_bednight[0][0]])
        return jsonify({"message": "Data successfully received", "data": data}), 200
    else:
        c.execute(
            '''
            INSERT INTO bednights (user_id, date, zero_six, six_twelve, twelve_eighteen, adults)
            VALUES (%s, %s, %s, %s, %s, %s)
            ''',
            (session['parent_id'], date, zero_six, six_twelve, twelve_eighteen, adults)
        )
        dbe.commit()


        return jsonify({"message": "Data successfully received", "data": data}), 200


        
@app.route('/manage_bednights/<start_date>/<end_date>', methods=['GET'])
def manage_bednights(start_date, end_date):
    if not auth():
        return redirect(url_for("login_page"))

    dbe, c = db.connect()

    # Αναμένουμε ότι bednights έχει πεδίο `date` σε μορφή DATE ή DATETIME
    c.execute(
        '''
        SELECT * FROM bednights
        WHERE user_id = %s AND date BETWEEN %s AND %s
        ORDER BY date
        ''',
        (session['parent_id'], start_date, end_date)
    )

    bednights_raw = c.fetchall()

    return render_template(
        'manage_bednights.html',
        title='Manage Bednights',
        bednights=bednights_raw,
        owner= session['user_hotel_owner'],
    )


@app.route('/update_bednight', methods=['POST'])
def update_bednight():
    if not auth():
        return redirect(url_for("login_page"))

    data = request.get_json()
    ide = data['id']
    field = data['field']
    old_value = data['old_value']
    new_value = data['new_value']

    # Προστασία: επέτρεψε μόνο συγκεκριμένα ονόματα πεδίων
    allowed_fields = {'zero_six', 'six_twelve', 'twelve_eighteen', 'adults'}
    if field not in allowed_fields:
        return jsonify({"error": "Invalid field name"}), 400

    conn, c = db.connect()

    # Safe string interpolation for column name (MUST be sanitized like above!)
    query = f'UPDATE bednights SET {field} = %s WHERE id = %s'
    c.execute(query, (new_value, ide))
    conn.commit()
    c.close()
    conn.close()

    print(f"Updating ID {ide}: {field} from {old_value} to {new_value}")
    return jsonify(success=True)

@app.route('/monthly', methods=['GET', 'POST'])
def select_monthly():
    
    if not auth():
        return redirect(url_for("login_page"))
    
    dbe,c = db.connect()
    c.execute('SELECT week FROM schedules WHERE user_id=%s', [session['user_id']])
    all_months = c.fetchall()
    month_dict = {
        '01' : 'january',
        '02' : 'february',
        '03' : 'march',
        '04' : 'april',
        '05' : 'may',
        '06' : 'june',
        '07' : 'july',
        '08' : 'august',
        '09' : 'september',
        '10' : 'october',
        '11' : 'november',
        '12' : 'december'
    }
    months=[]
    for month in all_months:
        month = month[0]
        date_str = month[1:][:-1]
        dates = date_str.split(' - ')
        for d in dates:
            m = d.split('-')[1] 
            print(m)
            if [f'{m}+{d.split("-")[0]}',f'{month_dict[m].capitalize()} {d.split("-")[0]}'] not in months:
                months.append([f'{m}+{d.split("-")[0]}',f'{month_dict[m].capitalize()} {d.split("-")[0]}'])
        
    
    
    return render_template('select_monthly.html', months=months)
def calculate_shift_hours(shift: str) -> int:
    try:
        start_str, end_str = shift.split('-')

        # Special handling for '24:00'
        if end_str.strip() == '24:00':
            end_str = '00:00'
            next_day = True
        else:
            next_day = False

        start_dt = datetime.strptime(start_str.strip(), '%H:%M')
        end_dt = datetime.strptime(end_str.strip(), '%H:%M')

        if next_day or end_dt <= start_dt:
            end_dt += timedelta(days=1)

        time_diff = end_dt - start_dt
        return int(time_diff.total_seconds() // 3600)

    except ValueError as e:
        raise ValueError(f"Invalid shift format '{shift}'. Expected 'HH:MM-HH:MM'. Error: {e}")




@app.route('/monthly_sched')
def monthly_sched():
    if not auth():
        return redirect(url_for("login_page"))

    # Step 1: Parse month and year
    month_str, year_str = request.args.get('month').split('+')
    month = int(month_str)
    year = int(year_str)

    first_day = datetime(year, month, 1).date()
    _, last_day_num = calendar.monthrange(year, month)
    last_day = datetime(year, month, last_day_num).date()

    # Step 2: DB Connection
    dbe, cursor = db.connect()

    # Get employee name mapping
    cursor.execute("SELECT id, name FROM employees")
    id_to_name = {int(emp_id): name for emp_id, name in cursor.fetchall()}

    # Get schedule records
    cursor.execute("SELECT week, schedule_json FROM schedules WHERE user_id=%s", [session['user_id']])
    rows = cursor.fetchall()
    dbe.close()

    # Step 3: Prepare schedule
    days_list = [(first_day + timedelta(days=i)) for i in range((last_day - first_day).days + 1)]
    schedule_per_employee = defaultdict(lambda: {d.isoformat(): "" for d in days_list})

    for week_range, sched_json in rows:
        try:
            week_start_str, week_end_str = week_range.split(' - ')
            week_start = datetime.strptime(week_start_str.strip('() '), "%Y-%m-%d").date()
            week_end = datetime.strptime(week_end_str.strip('() '), "%Y-%m-%d").date()

            if week_start <= last_day and week_end >= first_day:
                data = json.loads(sched_json.replace("'", '"'))

                for department, areas in data.items():
                    for section, shifts in areas.items():
                        for emp_id, day_data_list in shifts:
                            cur_emp_id = emp_id
                            emp_id = int(emp_id)
                            emp_name = id_to_name.get(emp_id, f"Unknown ({emp_id})")
                            # print(db.Employee(emp_id))
                            emp_code = db.Employee(emp_id, True).code
                            if emp_obj := db.Employee(emp_id, True).archived:
                                emp_name = f"{emp_name} (Archived)"
                                continue

                            for day_data in day_data_list:
                                for weekday, (shift, _) in day_data.items():
                                    day_index = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday'].index(weekday)
                                    shift_date = week_start + timedelta(days=day_index)

                                    if first_day <= shift_date <= last_day and shift:
                                        date_key = shift_date.isoformat()
                                        # X if it contains digits or the word "IN"
                                        # 5hrs = 0.5 else X
                                        entry = 'BRBR'
                                        if any(char.isdigit() for char in shift) or "IN" in shift.upper():
                                            if 'IN' not in shift.upper():
                                                # print(shift, type(shift), calculate_shift_hours(shift))
                                                if calculate_shift_hours(shift) == 5:
                                                    entry = "0.5"
                                                else:
                                                    entry = "X"
                                            else:
                                                entry = "X"
                                        else:
                                            entry = shift.split(' - ')[0].strip().upper()
                                            if entry == 'ARMY':
                                                entry = 'AR'
                                        existing = schedule_per_employee[(emp_name, emp_code)][date_key]
                                        schedule_per_employee[(emp_name, emp_code)][date_key] = (
                                            f"{existing} / {entry}" if existing else entry
                                        )

        except Exception as e:
            print(f"Error processing id {cur_emp_id} {week_range}: {e}")
    print('len: ', len(schedule_per_employee))
    return render_template(
        "view_monthly.html",
        month=f"{month:02d}",
        year=year,
        days=[d.isoformat() for d in days_list],
        data=schedule_per_employee
    )

import io 

@app.route('/download-monthly', methods=['POST'])
def download_schedule():
    data = request.get_json()

    df = pd.DataFrame.from_dict(data, orient='index')

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule')

    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=schedule.xlsx"
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return response
